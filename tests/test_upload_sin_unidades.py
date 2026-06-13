"""Tests del flag permitir_sin_unidades en POST /proyectos/{id}/unidades/excel/upload.

Caso real (Trinidad III, Maestra): proyecto agotado de deptos que solo conserva
estac/bodega. El Excel JB trae UNIDAD vacío. Por defecto el endpoint rechaza
(400) — protege contra scrapes vacíos por throttle que borrarían deptos. Con
permitir_sin_unidades=true, el scraper (que ya validó cross-source que el 0 es
real) puede sincronizar estac/bodega SIN dar de baja deptos.
"""
import io

from openpyxl import Workbook

from app.models import Proyecto, Unidad


def _jb_excel(unidad_rows_vacio=True):
    """Excel JB mínimo: UNIDAD (opcionalmente vacío) + ESTAC + BODEGA + INSTRUCCIONES."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("INSTRUCCIONES")
    # UNIDAD: fila0 = REQ/OPC, fila1 = labels. Sin filas de datos → vacío.
    u = wb.create_sheet("UNIDAD")
    u.append(["REQ", "REQ", "REQ"])
    u.append(["Unidad Número", "Modelo", "ValorUF"])
    if not unidad_rows_vacio:
        u.append(["101", "2D1B", 2500])
    # ESTACIONAMIENTOS
    e = wb.create_sheet("ESTACIONAMIENTOS")
    e.append(["OPC", "OPC"])
    e.append(["Número", "PrecioUF"])
    e.append(["55", 300])
    e.append(["98", 320])
    # BODEGAS
    b = wb.create_sheet("BODEGAS")
    b.append(["OPC", "OPC"])
    b.append(["Número", "PrecioUF"])
    b.append(["39", 100])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _crear_proy(db, pid, con_depto=False):
    if not db.query(Proyecto).filter(Proyecto.id == pid).first():
        db.add(Proyecto(id=pid, nombre=pid, extra={}))
        db.commit()
    if con_depto:
        if not db.query(Unidad).filter(Unidad.proyecto_id == pid, Unidad.numero == "101").first():
            db.add(Unidad(id=f"{pid}-101", proyecto_id=pid, numero="101", tipo="Depto",
                          modelo="2D1B", disponible=True))
            db.commit()


def test_unidad_vacia_sin_flag_da_400(client, admin, db):
    _, h = admin
    pid = "test-sin-unid-400"
    _crear_proy(db, pid)
    r = client.post(
        f"/proyectos/{pid}/unidades/excel/upload",
        headers=h,
        files={"file": ("x.xlsx", _jb_excel(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400
    assert "UNIDAD" in r.text


def test_unidad_vacia_con_flag_sincroniza_estac_bodega(client, admin, db):
    _, h = admin
    pid = "test-sin-unid-ok"
    _crear_proy(db, pid)
    r = client.post(
        f"/proyectos/{pid}/unidades/excel/upload?permitir_sin_unidades=true",
        headers=h,
        files={"file": ("x.xlsx", _jb_excel(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["inserted"] == 0
    assert body["dados_de_baja"] == []
    # estac/bodega quedaron en extra
    proy = db.query(Proyecto).filter(Proyecto.id == pid).first()
    db.refresh(proy)
    nums_est = {e["numero"] for e in (proy.extra or {}).get("estacionamientos", [])}
    nums_bod = {b["numero"] for b in (proy.extra or {}).get("bodegas", [])}
    assert {"55", "98"} <= nums_est
    assert {"39"} <= nums_bod


def test_flag_no_da_de_baja_deptos_existentes(client, admin, db):
    """SEGURIDAD: un upload sin deptos + flag NO debe dar de baja deptos existentes."""
    _, h = admin
    pid = "test-sin-unid-safe"
    _crear_proy(db, pid, con_depto=True)
    r = client.post(
        f"/proyectos/{pid}/unidades/excel/upload?permitir_sin_unidades=true",
        headers=h,
        files={"file": ("x.xlsx", _jb_excel(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    assert r.json()["dados_de_baja"] == []
    depto = db.query(Unidad).filter(Unidad.proyecto_id == pid, Unidad.numero == "101").first()
    db.refresh(depto)
    assert depto.disponible is True  # NO se dio de baja
