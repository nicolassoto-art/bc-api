"""Round-trip: JBImporter.build_jb_style_excel() -> app.routes.unidades._parse_jb_excel().

Confirma que el Excel sintético que arma el sync de marketplace/workview
(para reusar el endpoint /excel/upload, con upsert+baja seguro) es realmente
parseable por el lado bc-api -- sin esto, el sync diario fallaría en
silencio o el endpoint lo rechazaría por "formato no reconocido".
"""
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from app.routes.unidades import _is_jb_excel, _parse_jb_excel
from app.services.jb_importer import JBImporter


def _imp() -> JBImporter:
    return JBImporter(
        jb_email="x@example.com",
        jb_password="pw",
        bc_api_base="http://example.invalid",
        bc_jwt="dummy",
        imports_dir=Path(tempfile.mkdtemp()),
    )


def _unidad(numero, modelo="B", orientacion="SO", precio=3034.0, descuento=10.0, bono_pie=10.0):
    return {
        "numero": numero, "modelo": modelo, "tipologia": "1D - 1B", "tipo": "Depto",
        "orientacion": orientacion,
        "sup_total": 39.56, "sup_interior": 32.92, "sup_terraza": 6.64,
        "sup_logia": 0, "sup_jardin": 0,
        "precio_lista_uf": precio, "descuento_pct": descuento, "bono_pie_pct": bono_pie,
        "precio_final_uf": precio,
        "estac_flag": "optional", "bodega_flag": "optional", "pack_flag": "never",
        "disponible": True,
    }


def test_excel_generado_es_reconocido_como_formato_jb():
    imp = _imp()
    xlsx_path = imp.build_jb_style_excel("TEST01", [_unidad("1402")])
    wb = load_workbook(xlsx_path)
    assert _is_jb_excel(wb)


def test_roundtrip_una_unidad():
    imp = _imp()
    xlsx_path = imp.build_jb_style_excel("TEST01", [_unidad("1402", modelo="B", precio=3034.0)])
    wb = load_workbook(xlsx_path, data_only=True)
    rows, errors = _parse_jb_excel(wb)
    assert errors == []
    assert len(rows) == 1
    r = rows[0]
    assert r["numero_depto"] == "1402"
    assert r["modelo"] == "B"
    assert r["orientacion"] == "SO"
    assert r["precio_lista_uf"] == 3034.0
    assert r["sup_total"] == 39.56


def test_roundtrip_multiples_unidades_preserva_todas():
    imp = _imp()
    unidades = [_unidad("1402"), _unidad("2111", modelo="K"), _unidad("2512", modelo="L")]
    xlsx_path = imp.build_jb_style_excel("TEST01", unidades)
    wb = load_workbook(xlsx_path, data_only=True)
    rows, errors = _parse_jb_excel(wb)
    assert errors == []
    assert {r["numero_depto"] for r in rows} == {"1402", "2111", "2512"}


def test_excel_vacio_sin_unidades_sigue_siendo_formato_jb_valido():
    imp = _imp()
    xlsx_path = imp.build_jb_style_excel("TEST01", [])
    wb = load_workbook(xlsx_path)
    assert _is_jb_excel(wb)
    rows, errors = _parse_jb_excel(load_workbook(xlsx_path, data_only=True))
    assert rows == []
