"""
diag_scrape_live.py — Diagnóstico SOLO-LECTURA de las 3 fuentes de unidades del scraper.

NO escribe en bc-api. NO borra nada. NO llama run()/wipe.
Reutiliza los primitivos de JBImporter (login, _jb_httpx, _click_tab, _scrape_table_rows)
para ver EXACTAMENTE qué devuelve JB ahora en los proyectos vaciados.

Para cada proyecto mide:
  1. API  → status HTTP de cada endpoint de unidades + conteo (detecta 429/403/empty)
  2. DOM  → trayectoria de filas en la tabla Unidades a lo largo de ~20s
            (distingue render-lento de tabla-genuinamente-vacía)
  3. Excel→ descarga "plantilla vacía" Y "prerellena con datos" por separado,
            cuenta filas REALES de cada una (no solo bytes)

Veredicto por proyecto:
  - API_OK            → JB tiene la data por API (el problema es DOM/timing en el importer)
  - RATE_LIMITED      → algún endpoint devolvió 429/403 (bloqueo temporal)
  - JB_EMPTY          → JB realmente no publica stock (API 200 + 0, tabla 0, excel 0 filas)
  - DOM_SLOW_RENDER   → tabla parte en 0 pero se llena con el tiempo (subir el wait)

Env: JETBROKERS_EMAIL, JETBROKERS_PASS, BC_API_JWT (dummy ok, no se usa para escribir)
     DIAG_IDS (opcional, csv) — default: igfwzvh2,g3jwrroe,iaiq9ith
"""
from __future__ import annotations
import asyncio, json, os, sys, time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from app.services.jb_importer import JBImporter  # noqa: E402

TEST_IDS = [s.strip() for s in os.environ.get("DIAG_IDS", "igfwzvh2,g3jwrroe,iaiq9ith").split(",") if s.strip()]

UNIT_ENDPOINTS = [
    "/store/project/{id}/available",
    "/projects/{id}/units",
    "/projects/{id}/apartments",
    "/units?projectId={id}",
    "/apartments?projectId={id}",
]


def _items_of(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for k in ("data", "elements", "units", "apartments"):
            v = data.get(k)
            if isinstance(v, list):
                return v
    return []


def _deptos(items):
    return sum(1 for u in items if isinstance(u, dict)
               and isinstance(u.get("apartmentModel"), dict)
               and u["apartmentModel"].get("name"))


async def probe_api(imp, jb_id) -> dict:
    """Hit raw a cada endpoint, captura status + conteo. Revela 429/403/empty."""
    res = {"endpoints": [], "best_units": 0, "best_deptos": 0, "any_429": False, "any_403": False}
    async with imp._jb_httpx() as cli:
        # proyecto base
        try:
            rp = await cli.get(f"/projects/{jb_id}")
            res["project_status"] = rp.status_code
        except Exception as e:
            res["project_status"] = f"ERR:{e}"
        for tmpl in UNIT_ENDPOINTS:
            ep = tmpl.replace("{id}", jb_id)
            row = {"ep": ep, "status": None, "items": 0, "deptos": 0}
            try:
                r = await cli.get(ep)
                row["status"] = r.status_code
                if r.status_code == 429:
                    res["any_429"] = True
                if r.status_code == 403:
                    res["any_403"] = True
                if r.status_code == 200:
                    items = _items_of(r.json())
                    row["items"] = len(items)
                    row["deptos"] = _deptos(items)
                    res["best_units"] = max(res["best_units"], len(items))
                    res["best_deptos"] = max(res["best_deptos"], row["deptos"])
            except Exception as e:
                row["status"] = f"ERR:{str(e)[:40]}"
            res["endpoints"].append(row)
            await asyncio.sleep(0.3)
    return res


async def probe_dom(imp, jb_id) -> dict:
    """Abre editor, va a Unidades, mide trayectoria de filas tbody en el tiempo."""
    res = {"row_trajectory": [], "max_rows": 0, "general_fields": 0}
    page = imp._page
    try:
        await page.goto(f"https://app.jetbrokers.io/projects/edit/{jb_id}",
                        wait_until="networkidle", timeout=60_000)
        await page.wait_for_timeout(4_000)
        await imp._dismiss_popups()
        # ¿el form general cargó? (control de salud de la sesión)
        res["general_fields"] = await page.evaluate(
            "() => document.querySelectorAll('input,mat-select,textarea').length")
        # ir a Unidades
        await imp._click_tab("Unidades")
        # medir filas cada 2s por ~24s SIN scrollear (queremos ver render natural)
        for t in range(12):
            await page.wait_for_timeout(2_000)
            n = await page.evaluate("""() => {
                const scope = document.querySelector('mat-tab-body.mat-mdc-tab-body-active') || document.body;
                return scope.querySelectorAll('table tbody tr').length;
            }""")
            res["row_trajectory"].append({"t_s": (t + 1) * 2, "rows": n})
            res["max_rows"] = max(res["max_rows"], n)
        # ¿hay mensaje de "sin datos"/spinner?
        res["empty_state"] = await page.evaluate(r"""() => {
            const scope = document.querySelector('mat-tab-body.mat-mdc-tab-body-active') || document.body;
            const txt = (scope.innerText || '').toLowerCase();
            return {
                no_data: /no hay|sin datos|sin resultados|no se encontr|empty|vac[ií]o/.test(txt),
                spinner: !!scope.querySelector('mat-spinner, mat-progress-spinner, .spinner, .loading'),
                text_sample: (scope.innerText || '').slice(0, 200),
            };
        }""")
    except Exception as e:
        res["error"] = str(e)[:200]
    return res


def _count_xlsx_rows(path: Path) -> dict:
    """Cuenta filas con contenido por hoja (revela plantilla vacía vs con datos)."""
    import openpyxl
    out = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"error": str(e)[:120]}
    for ws in wb.worksheets:
        data_rows = 0
        sample = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [c for c in row if c not in (None, "")]
            if not cells:
                continue
            # saltar filas de encabezado/instrucción obvias contando solo desde la 2da con contenido
            if i >= 1:
                data_rows += 1
                if len(sample) < 2:
                    sample.append([str(c)[:18] for c in row[:6]])
        out[ws.title] = {"max_row": ws.max_row, "content_rows": data_rows, "sample": sample}
    wb.close()
    return out


async def probe_excel(imp, jb_id) -> dict:
    """Descarga plantilla vacía y prerellena por separado; cuenta filas reales."""
    res = {}
    page = imp._page
    outdir = imp.imports_dir / jb_id
    outdir.mkdir(parents=True, exist_ok=True)
    try:
        await imp._click_tab("Stock")
        await page.wait_for_timeout(2_500)
    except Exception as e:
        res["stock_tab_error"] = str(e)[:120]
        return res
    variants = {
        "prerellena": ["prerellena con datos", "prerellena", "con datos"],
        "vacia": ["Descargar plantilla", "Descargar"],
    }
    for label, texts in variants.items():
        got = False
        for btn in texts:
            try:
                async with page.expect_download(timeout=12_000) as dl:
                    await page.locator(
                        f"button:has-text('{btn}'), a:has-text('{btn}')").first.click(timeout=4_000)
                d = await dl.value
                p = outdir / f"diag-{label}.xlsx"
                await d.save_as(str(p))
                if p.exists() and p.stat().st_size > 256:
                    res[label] = {"btn": btn, "size": p.stat().st_size, "sheets": _count_xlsx_rows(p)}
                    got = True
                    break
            except Exception:
                continue
        if not got:
            res[label] = {"btn": None, "note": "no se pudo descargar esta variante"}
    return res


def verdict(api, dom, excel) -> str:
    if api.get("any_429") or api.get("any_403"):
        return "RATE_LIMITED (429/403 en endpoints de stock)"
    if api.get("best_deptos", 0) > 0 or api.get("best_units", 0) > 0:
        return f"API_OK ({api['best_units']} units / {api['best_deptos']} deptos) → el importer debería usar API, problema es DOM/orden"
    # excel prerellena con filas?
    pre = (excel.get("prerellena") or {}).get("sheets") or {}
    pre_unidad = max([v.get("content_rows", 0) for k, v in pre.items()
                      if isinstance(v, dict) and "unidad" in k.lower()] or [0])
    if pre_unidad > 0:
        return f"EXCEL_OK (prerellena UNIDAD={pre_unidad} filas) → usar Excel prerellena como fuente"
    traj = dom.get("row_trajectory") or []
    if dom.get("max_rows", 0) > 0:
        first = traj[0]["rows"] if traj else 0
        if first == 0:
            return f"DOM_SLOW_RENDER (0→{dom['max_rows']} filas) → subir wait antes de leer"
        return f"DOM_OK ({dom['max_rows']} filas)"
    return "JB_EMPTY (API 200+0, excel 0 filas, DOM 0 filas) → JB realmente no publica stock ahora"


async def main():
    imp = JBImporter(
        jb_email=os.environ["JETBROKERS_EMAIL"],
        jb_password=os.environ["JETBROKERS_PASS"],
        bc_api_base=os.environ.get("BC_API_BASE", "https://bc-api.178-105-91-29.nip.io"),
        bc_jwt=os.environ.get("BC_API_JWT", "dummy-no-write"),
        imports_dir=Path("imports/_diag"),
    )
    await imp.login()
    results = {}
    try:
        for jb_id in TEST_IDS:
            print(f"\n{'='*70}\n=== {jb_id} ===\n{'='*70}", flush=True)
            api = await probe_api(imp, jb_id)
            print("  [API]", flush=True)
            print(f"    project status: {api.get('project_status')}", flush=True)
            for r in api["endpoints"]:
                print(f"    {r['status']!s:>6}  items={r['items']:<4} deptos={r['deptos']:<4} {r['ep']}", flush=True)
            dom = await probe_dom(imp, jb_id)
            print("  [DOM]", flush=True)
            print(f"    general inputs visibles: {dom.get('general_fields')}", flush=True)
            traj = " ".join(f"{p['t_s']}s:{p['rows']}" for p in dom.get("row_trajectory", []))
            print(f"    filas tbody en el tiempo: {traj}", flush=True)
            es = dom.get("empty_state") or {}
            print(f"    empty_state: no_data={es.get('no_data')} spinner={es.get('spinner')}", flush=True)
            if es.get("text_sample"):
                print(f"    muestra texto: {es['text_sample']!r}", flush=True)
            excel = await probe_excel(imp, jb_id)
            print("  [EXCEL]", flush=True)
            for variant in ("prerellena", "vacia"):
                v = excel.get(variant) or {}
                if v.get("sheets"):
                    sheet_summary = {k: s.get("content_rows") for k, s in v["sheets"].items() if isinstance(s, dict)}
                    print(f"    {variant}: btn={v.get('btn')!r} size={v.get('size')} filas_por_hoja={sheet_summary}", flush=True)
                else:
                    print(f"    {variant}: {v.get('note', 'sin datos')}", flush=True)
            v = verdict(api, dom, excel)
            print(f"  ➜ VEREDICTO: {v}", flush=True)
            results[jb_id] = {"api": api, "dom": dom, "excel": excel, "verdict": v}
            await asyncio.sleep(2)
    finally:
        await imp.close()

    print(f"\n{'='*70}\n=== RESUMEN ===\n{'='*70}", flush=True)
    for jb_id, r in results.items():
        print(f"  {jb_id:<12} {r['verdict']}", flush=True)
    Path("imports/_diag").mkdir(parents=True, exist_ok=True)
    Path("imports/_diag/scrape_diag.json").write_text(json.dumps(results, indent=2, default=str))


if __name__ == "__main__":
    asyncio.run(main())
