"""
sync_jb_stock.py — Sync liviano de stock JetBrokers → bc-api (bajo consumo JB).

A diferencia de import_jb.py (full wipe+reimport: notas comerciales, condiciones
comerciales, fotos, planos), este script SOLO entra a la tab "Stock" del editor JB
(JBImporter.sync_stock_light), descarga el Excel y lo sube por /excel/upload
(upsert+baja ya seguro del lado bc-api, mismo endpoint que usa MNK). No wipea el
proyecto, no toca extra.notas_html/extra.comercial/fotos/planos.

Antes de subir, aborta si el Excel implica una baja masiva de deptos (>=8 absolutos
Y >=40% de los disponibles actuales) — el endpoint /excel/upload de bc-api NO tiene
esta protección server-side (confirmado leyendo app/routes/unidades.py). Replica el
patrón ya probado en MNK (ver CLAUDE.md del scraper MNK, "Guarda anti-baja-masiva").

Uso:
  python3 scripts/sync_jb_stock.py <jb_id>
  python3 scripts/sync_jb_stock.py <jb_id> --headed   # debug local, browser visible

Env requeridas: JETBROKERS_EMAIL, JETBROKERS_PASS, BC_API_JWT (o BC_TOKEN)
Opcional: BC_API_BASE (default https://bc-api.178-105-91-29.nip.io)
Requiere: proyecto ya importado (extra.jb_id) — correr import_jb.py primero si no existe.
"""
from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import httpx
from openpyxl import load_workbook

from app.services.jb_importer import JBImporter
from app.routes.unidades import _is_jb_excel, _parse_jb_excel


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("sync_jb_stock")

BAJA_MIN_ABS = 8
BAJA_MIN_PCT = 0.40


async def get_jwt(bc_api_base: str) -> str:
    """Obtiene un JWT — prioriza BC_API_JWT (ya fresco), sino exchange con BC_TOKEN."""
    jwt = os.environ.get("BC_API_JWT")
    if jwt:
        return jwt
    bc_token = os.environ.get("BC_TOKEN")
    if not bc_token:
        raise RuntimeError("Faltan BC_API_JWT o BC_TOKEN en env")
    log.info("🔑 Exchange BC_TOKEN → JWT...")
    async with httpx.AsyncClient(timeout=15.0) as cli:
        r = await cli.post(f"{bc_api_base}/auth/exchange", json={"bc_token": bc_token})
    if not r.is_success:
        raise RuntimeError(f"/auth/exchange falló: HTTP {r.status_code}: {r.text[:200]}")
    data = r.json()
    return data.get("access_token") or data.get("token")


def _is_depto(u: dict) -> bool:
    t = (u.get("tipo") or "").lower()
    return t in ("depto", "departamento", "apartment", "")


def check_baja_masiva(excel_numeros: set[str], unidades_actuales: list[dict]) -> tuple[bool, str]:
    """Compara deptos disponibles actuales en bc-api vs los que trae el Excel nuevo.

    Devuelve (abortar, motivo). Replica el patrón probado de MNK
    (stock_interno_uploader.py) porque /excel/upload de bc-api no trae esta guardia.
    """
    disponibles_actuales = [u for u in unidades_actuales if _is_depto(u) and u.get("disponible")]
    if not disponibles_actuales:
        return False, "sin deptos disponibles previos — nada que resguardar"
    numeros_actuales = {str(u.get("numero") or "").strip() for u in disponibles_actuales}
    se_darian_de_baja = numeros_actuales - excel_numeros
    n_baja = len(se_darian_de_baja)
    n_total = len(numeros_actuales)
    pct_baja = (n_baja / n_total) if n_total else 0
    if n_baja >= BAJA_MIN_ABS and pct_baja >= BAJA_MIN_PCT:
        return True, (
            f"BAJA MASIVA: {n_baja}/{n_total} deptos disponibles ({pct_baja:.0%}) "
            f"no aparecen en el Excel nuevo — abortando sin subir. "
            f"Deptos: {sorted(se_darian_de_baja)[:15]}"
        )
    return False, f"delta OK: {n_baja}/{n_total} bajas ({pct_baja:.0%})"


async def run_one(jb_id: str, bc_base: str, jwt: str, headless: bool = True) -> int:
    imp = JBImporter(
        jb_email=os.environ["JETBROKERS_EMAIL"],
        jb_password=os.environ["JETBROKERS_PASS"],
        bc_api_base=bc_base,
        bc_jwt=jwt,
        imports_dir=Path(os.environ.get("IMPORTS_DIR", "imports")),
        headless=headless,
    )
    try:
        await imp.login()

        current = await imp.find_proyecto_by_jb_id(jb_id)
        if not current:
            log.error(f"✗ No existe proyecto con extra.jb_id={jb_id} en bc-api — correr import_jb.py primero (import completo)")
            return 1
        proyecto_id = current["id"]
        log.info(f"▶ {current.get('nombre')} (proyecto_id={proyecto_id})")

        xlsx_path = await imp.sync_stock_light(jb_id)
        if not xlsx_path:
            log.error("✗ No se pudo descargar el Excel de stock (botón no encontrado)")
            return 1

        wb = load_workbook(xlsx_path, data_only=True)
        if not _is_jb_excel(wb):
            log.error(f"✗ Excel descargado no tiene el formato JB esperado (sheets: {wb.sheetnames})")
            return 1
        rows, parse_errors = _parse_jb_excel(wb)
        if parse_errors:
            log.warning(f"   parse warnings: {parse_errors[:5]}")
        excel_numeros = {str(d.get("numero_depto") or "").strip() for d in rows if d.get("numero_depto")}
        log.info(f"   📋 Excel trae {len(excel_numeros)} deptos")

        unidades_actuales = current.get("unidades") or []
        abortar, motivo = check_baja_masiva(excel_numeros, unidades_actuales)
        log.info(f"   🛡  guardia anti-baja: {motivo}")
        if abortar:
            log.error(f"✗ {motivo}")
            return 1

        result = await imp.upload_jb_excel(proyecto_id, xlsx_path)
        if result.get("status") == "error":
            log.error(f"✗ upload falló: {result}")
            return 1
        log.info(
            f"   ✓ upload: {result.get('inserted', 0)} ins, {result.get('updated', 0)} upd, "
            f"{len(result.get('errors', []))} err"
        )

        live_check = await imp.verify_live(proyecto_id, {})
        log.info(f"   verify_live: {live_check.get('live', {}).get('unidades')} unidades totales en bc-api")

        print(json.dumps({
            "jb_id": jb_id,
            "proyecto_id": proyecto_id,
            "excel_deptos": len(excel_numeros),
            "upload": result,
        }, indent=2, default=str))
        return 0
    finally:
        await imp.close()


def main():
    parser = argparse.ArgumentParser(description="Sync liviano de stock JB → bc-api (sin wipe, sin notas/fotos)")
    parser.add_argument("jb_id", help="JB project ID (ej: Sfp8j2Sq)")
    parser.add_argument("--headed", action="store_true", help="Playwright con browser visible (debug local)")
    args = parser.parse_args()

    bc_base = os.environ.get("BC_API_BASE", "https://bc-api.178-105-91-29.nip.io").rstrip("/")

    async def _main():
        jwt = await get_jwt(bc_base)
        code = await run_one(args.jb_id, bc_base, jwt, headless=not args.headed)
        sys.exit(code)

    asyncio.run(_main())


if __name__ == "__main__":
    main()
