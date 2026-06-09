"""Unidades de un proyecto: CRUD + upload/download Excel."""
from typing import List
import io
import uuid
from datetime import datetime
from fastapi import APIRouter, BackgroundTasks, Body, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..db import get_db
from ..deps.auth import stock_access
from ..models import Proyecto, Unidad, Usuario
from ..schemas import UnidadIn, UnidadOut
from ..services import email_service

router = APIRouter(prefix="/proyectos/{proyecto_id}/unidades", tags=["unidades"])

HEADERS = [
    "numero_depto", "modelo", "tipologia", "tipo", "orientacion",
    "sup_total", "sup_interior", "sup_terraza", "sup_logia", "sup_jardin",
    "precio_lista_uf", "descuento_pct", "bono_pie_pct", "precio_final_uf",
    "estac_flag", "bodega_flag", "pack_flag", "disponible",
]

# ── JetBrokers Excel format support ──────────────────────────────────────────
# JB Excel v2.4 tiene 4 sheets: INSTRUCCIONES, UNIDAD, ESTACIONAMIENTOS, BODEGAS
# En sheet UNIDAD: fila 1 = "REQ" markers, fila 2 = labels reales, fila 3+ = data
JB_SHEETS = {"INSTRUCCIONES", "UNIDAD", "ESTACIONAMIENTOS", "BODEGAS"}

# Mapeo de label JB (en sheet UNIDAD fila 2) → key interna bc-api
JB_UNIDAD_MAP = {
    "Unidad Número":            "numero_depto",
    "Dormitorios":              "_dormitorios",
    "Baños":                    "_banos",
    "Modelo":                   "modelo",
    "Orientacion":              "orientacion",
    "Sup Interior":             "sup_interior",
    "Sup Terraza":              "sup_terraza",
    "Sup Logia":                "sup_logia",
    "Sup Jardin":               "sup_jardin",
    "Sup Total":                "sup_total",
    "ValorUF":                  "precio_lista_uf",
    "Descuento":                "descuento_pct",
    "Bonopie":                  "bono_pie_pct",
    "Cotiza Estacionamiento":   "estac_flag",
    "Cotiza Bodega":            "bodega_flag",
    "Cotiza Pack":              "pack_flag",
    "Estacionamiento Número":   "_estac_num",
    "Bodega Número":            "_bodega_num",
    "Pack Número":              "_pack_num",
    "Id Externo":               "_jb_id",
    "_ModeloWarning":           "_modelo_warning",  # ext. Maestra: depto con modelo no registrado
}

# Mapeo de valores "Cotiza X" JB → flag bc-api
JB_COTIZA_MAP = {
    "obligatorio": "required",
    "opcional":    "optional",
    "nunca":       "never",
    "required":    "required",
    "optional":    "optional",
    "never":       "never",
}


def _is_jb_excel(wb) -> bool:
    """Detecta si el .xlsx es formato JB (tiene los 4 sheets típicos)."""
    return JB_SHEETS.issubset(set(wb.sheetnames))


def _normalize_label(s: str) -> str:
    """Normaliza header: sin tildes, lowercase, sin puntos, espacios normalizados.
    Sup./Sup./Superficie/SUP. → 'sup'; Logía/Logia → 'logia'; Jardín/Jardin → 'jardin'.
    """
    import unicodedata
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode("ascii")
    s = s.lower().replace(".", "").replace("º", "").replace("°", "")
    s = " ".join(s.split())  # colapsa espacios
    # equivalencias de prefijo "superficie" → "sup"
    if s.startswith("superficie "):
        s = "sup " + s[len("superficie "):]
    return s


def _build_idx_map(labels: list[str]) -> dict[int, str]:
    """Mapea índice de columna → bc-api key, con matching robusto (case+tildes+sinónimos)."""
    norm_to_key = {_normalize_label(k): v for k, v in JB_UNIDAD_MAP.items()}
    # Sinónimos extra (defensivo, por si JB cambia headers)
    extra = {
        "sup logia": "sup_logia", "sup logía": "sup_logia",
        "superficie logia": "sup_logia", "superficie logía": "sup_logia",
        "sup jardin": "sup_jardin", "sup jardín": "sup_jardin",
        "superficie jardin": "sup_jardin", "superficie jardín": "sup_jardin",
        "sup interior": "sup_interior", "superficie interior": "sup_interior",
        "sup terraza": "sup_terraza", "superfice terraza": "sup_terraza",  # typo JB conocido
        "superficie terraza": "sup_terraza",
        "sup total": "sup_total", "superficie total": "sup_total",
    }
    for k, v in extra.items():
        norm_to_key[_normalize_label(k)] = v
    idx_map: dict[int, str] = {}
    for i, label in enumerate(labels):
        norm = _normalize_label(label)
        if norm in norm_to_key:
            idx_map[i] = norm_to_key[norm]
    return idx_map


def _parse_jb_excel(wb) -> tuple[list[dict], list[str]]:
    """Parsea sheet UNIDAD del Excel JB → lista de dicts compatibles con bc-api.

    Retorna (rows_data, errors). Cada row ya tiene los nombres normalizados a bc-api.
    """
    import logging
    log = logging.getLogger(__name__)
    ws = wb["UNIDAD"]
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 3:
        return [], ["Sheet UNIDAD vacío"]
    # Fila 2 (index 1) tiene los labels reales
    labels = [str(c or "").strip() for c in all_rows[1]]
    idx_map = _build_idx_map(labels)
    # Loggear qué columnas se mapearon (debug) y cuáles quedaron sin mapear
    mapped_keys = sorted(set(idx_map.values()))
    unmapped = [labels[i] for i in range(len(labels)) if i not in idx_map and labels[i]]
    log.info(f"   📋 Excel UNIDAD mapeado: {mapped_keys}")
    if unmapped:
        log.info(f"   📋 Headers sin mapear: {unmapped}")
    out_rows = []
    errors = []
    for row_idx, row in enumerate(all_rows[2:], start=3):
        if not row or all(c is None or c == "" for c in row):
            continue
        d: dict = {}
        for i, val in enumerate(row):
            key = idx_map.get(i)
            if not key:
                continue
            d[key] = val
        num = str(d.get("numero_depto") or "").strip()
        if not num:
            errors.append(f"Fila {row_idx}: falta 'Unidad Número'")
            continue
        # Traducir cotiza_X a flags
        for k in ("estac_flag", "bodega_flag", "pack_flag"):
            v = d.get(k)
            if v is not None:
                d[k] = JB_COTIZA_MAP.get(str(v).strip().lower(), "optional")
        # Componer tipologia. Preferimos derivar del MODELO (preserva sufijos como
        # "g" de guardarropas o "M" de baño matrimonial), con fallback a
        # Dormitorios + Baños si el modelo no parsea.
        modelo_str = (d.get("modelo") or "").strip()
        tipologia_from_modelo = None
        if modelo_str:
            import re as _re_mod
            # Patrón Maestra: [N]D[g?][N]B[sufijo?]  · ej. 2D1B, 2Dg1B, 2D2BM, 1D1BS
            m_mod = _re_mod.match(r"^(\d+)D(g?)(\d+)B([A-Z]*)$", modelo_str, _re_mod.IGNORECASE)
            if m_mod:
                dorm_part = f"{m_mod.group(1)}D{m_mod.group(2)}"     # "2D" o "2Dg"
                bano_part = f"{m_mod.group(3)}B{m_mod.group(4)}"     # "1B" / "2BM" / "1BS"
                tipologia_from_modelo = f"{dorm_part} - {bano_part}"
        if tipologia_from_modelo:
            d["tipologia"] = tipologia_from_modelo
        elif d.get("_dormitorios") is not None and d.get("_banos") is not None:
            try:
                d["tipologia"] = f"{int(d['_dormitorios'])}D - {int(d['_banos'])}B"
            except Exception:
                pass
        elif modelo_str.lower().startswith("studio"):
            # Studio: PlanOk no da dormitorios → tipología "Estudio" (español)
            d["tipologia"] = "Estudio"
        # Regla (2026-06-08, pedido del usuario): SI la unidad APARECE en la
        # hoja UNIDAD del Excel, está DISPONIBLE — punto. El Excel del scraper
        # trae solo las unidades vigentes; el resto se asume vendido/no-disp.
        # IGNORAMOS la columna 'Disponible' del Excel JB (los humanos a veces
        # ponen 'FALSE' en una fila para marcar 'vendida', pero la fila ya
        # NO debería estar en el Excel → eso es un error del Excel, no del
        # sistema). Forzamos True acá; las bajas reales = ausencia en el Excel.
        # NOTA: estac/bodega tienen su propia lógica de pack (líneas 272, 297)
        # — eso queda intacto porque vive en parsers distintos.
        d["disponible"] = True
        # tipo: deducir
        d.setdefault("tipo", "Depto")
        out_rows.append(d)
    return out_rows, errors


def _parse_jb_estacionamientos(wb) -> list[dict]:
    """Parsea sheet ESTACIONAMIENTOS del Excel JB."""
    if "ESTACIONAMIENTOS" not in wb.sheetnames:
        return []
    ws = wb["ESTACIONAMIENTOS"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    labels = [str(c or "").strip() for c in rows[1]]
    out = []
    for r in rows[2:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        item = {labels[i]: r[i] for i in range(min(len(labels), len(r))) if labels[i]}
        if item.get("Número"):
            out.append(item)
    return out


def _parse_jb_bodegas(wb) -> list[dict]:
    """Parsea sheet BODEGAS del Excel JB."""
    if "BODEGAS" not in wb.sheetnames:
        return []
    ws = wb["BODEGAS"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    labels = [str(c or "").strip() for c in rows[1]]
    out = []
    for r in rows[2:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        item = {labels[i]: r[i] for i in range(min(len(labels), len(r))) if labels[i]}
        if item.get("Número"):
            out.append(item)
    return out


def _parse_jb_packs(wb) -> list[dict]:
    """Parsea hoja PACKS opcional del Excel JB (extensión Maestra).

    Cuando está presente, contiene la composición autoritativa de cada pack
    (numero, estacionamientos CSV, bodegas CSV, precio_uf). Esto reemplaza
    al método de agrupar por 'Pack Número' de las hojas individuales, que
    perdía info cuando una unidad debía aparecer en varios packs por
    declaraciones no recíprocas en el Excel de Maestra.
    """
    if "PACKS" not in wb.sheetnames:
        return []
    ws = wb["PACKS"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    labels = [str(c or "").strip() for c in rows[1]]
    out = []
    for r in rows[2:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        item = {labels[i]: r[i] for i in range(min(len(labels), len(r))) if labels[i]}
        if item.get("Pack Número"):
            out.append(item)
    return out


def _build_jb_extras(jb_estac: list[dict], jb_bodegas: list[dict], jb_packs: list[dict] | None = None) -> dict:
    """Construye los campos de extra que el frontend lee: estacionamientos, bodegas, packs.

    Reglas (alineadas con el scraper Maestra):
      - "Solo en pack": las unidades que aparecen en algún pack quedan disponible=false
        en sus listas individuales (siguen como ficha técnica, no se venden sueltas).
      - "Conflictos: crear ambos packs": cada agrupamiento por "Pack Número" genera 1 pack.
      - Conserva _estacionamientos_dom / _bodegas_dom como diagnóstico.

    Tenencia: campo opcional de las hojas del Excel JB extendidas; se mapea a cada item.
    """
    def _num(v):
        try:
            return float(v) if v not in (None, "") else None
        except (ValueError, TypeError):
            return None

    def _str(v):
        return str(v).strip() if v not in (None, "") else ""

    # Indexar por "Pack Número" → para armar packs
    packs_by_num: dict[str, dict] = {}

    estacs_out = []
    for i, e in enumerate(jb_estac or []):
        numero = _str(e.get("Número"))
        if not numero:
            continue
        pack_num = _str(e.get("Pack Número"))
        item = {
            "id": f"jb-est-{i}",
            "numero": numero,
            "precio_uf": _num(e.get("PrecioUF")),
            "nivel": _str(e.get("Nivel")),
            "tipo": _str(e.get("Tipo")),
            "tenencia": _str(e.get("Tenencia")),
            # "solo en pack": si la fila trae Pack Número, NO se vende suelto
            "disponible": not bool(pack_num),
        }
        if pack_num:
            item["pack_numero"] = pack_num
            grp = packs_by_num.setdefault(pack_num, {"estacionamientos": [], "bodegas": [], "precio_uf": 0.0})
            grp["estacionamientos"].append(numero)
            grp["precio_uf"] += item["precio_uf"] or 0
        estacs_out.append(item)

    bodegas_out = []
    for i, b in enumerate(jb_bodegas or []):
        numero = _str(b.get("Número"))
        if not numero:
            continue
        pack_num = _str(b.get("Pack Número"))
        # El campo "Tipo" de la hoja BODEGAS lleva la tenencia (Dominio / Uso y goce),
        # mismo criterio que en estacionamientos. "Tenencia" queda como fallback.
        tipo_v = _str(b.get("Tipo")) or _str(b.get("Tenencia"))
        item = {
            "id": f"jb-bod-{i}",
            "numero": numero,
            "precio_uf": _num(b.get("PrecioUF")),
            "superficie": _num(b.get("Superficie")),
            "tipo": tipo_v,
            "tenencia": _str(b.get("Tenencia")),  # compat retro
            "disponible": not bool(pack_num),
        }
        if pack_num:
            item["pack_numero"] = pack_num
            grp = packs_by_num.setdefault(pack_num, {"estacionamientos": [], "bodegas": [], "precio_uf": 0.0})
            grp["bodegas"].append(numero)
            grp["precio_uf"] += item["precio_uf"] or 0
        bodegas_out.append(item)

    # Construcción de packs: si el Excel trae hoja PACKS (extensión Maestra),
    # usar esa composición autoritativa. Si no, fallback a agrupar por Pack Número
    # de las hojas individuales (formato JB clásico, que pierde info en packs
    # no recíprocos).
    packs_out = []
    if jb_packs:
        for i, pk in enumerate(jb_packs):
            pn = _str(pk.get("Pack Número"))
            if not pn:
                continue
            estacs_raw = _str(pk.get("Estacionamientos"))
            bodegas_raw = _str(pk.get("Bodegas"))
            estacs_list = [s.strip() for s in estacs_raw.split(",") if s.strip()]
            bodegas_list = [s.strip() for s in bodegas_raw.split(",") if s.strip()]
            packs_out.append({
                "id": f"jb-pack-{i}",
                "numero": pn,
                "estacionamientos": estacs_list,
                "bodegas": bodegas_list,
                "precio_uf": _num(pk.get("PrecioUF")),
                "disponible": True,
            })
    else:
        for i, (pn, grp) in enumerate(sorted(packs_by_num.items())):
            packs_out.append({
                "id": f"jb-pack-{i}",
                "numero": pn,
                "estacionamientos": grp["estacionamientos"],
                "bodegas": grp["bodegas"],
                "precio_uf": grp["precio_uf"] or None,
                "disponible": True,
            })

    # Re-calcular disponibilidad de estac/bodegas basado en los packs reales:
    # una unidad solo queda disponible=false si aparece en algún pack VÁLIDO
    # (≥2 miembros). Si solo está en un pack huérfano (raro pero posible si
    # los datos están inconsistentes), la dejamos disponible.
    en_pack_estac = set()
    en_pack_bodega = set()
    for pk in packs_out:
        n_total = len(pk.get("estacionamientos") or []) + len(pk.get("bodegas") or [])
        if n_total < 2:
            continue  # pack huérfano, no marca a sus miembros como "solo en pack"
        for n in pk.get("estacionamientos") or []:
            en_pack_estac.add(n)
        for n in pk.get("bodegas") or []:
            en_pack_bodega.add(n)
    for it in estacs_out:
        it["disponible"] = it["numero"] not in en_pack_estac
        if it["numero"] in en_pack_estac:
            # Si no tenía pack_numero del fallback, busquemos al primer pack que lo incluya
            if not it.get("pack_numero"):
                for pk in packs_out:
                    if it["numero"] in (pk.get("estacionamientos") or []):
                        it["pack_numero"] = pk["numero"]
                        break
    for it in bodegas_out:
        it["disponible"] = it["numero"] not in en_pack_bodega
        if it["numero"] in en_pack_bodega:
            if not it.get("pack_numero"):
                for pk in packs_out:
                    if it["numero"] in (pk.get("bodegas") or []):
                        it["pack_numero"] = pk["numero"]
                        break

    # Filtrar packs huérfanos al final (no llegan al frontend)
    packs_out = [p for p in packs_out if (len(p.get("estacionamientos") or []) + len(p.get("bodegas") or [])) >= 2]

    return {
        "_excel_format": "jb_v2.4",
        "_estacionamientos_dom": jb_estac,  # raw, para diagnóstico
        "_bodegas_dom": jb_bodegas,         # raw, para diagnóstico
        "estacionamientos": estacs_out,     # shape que lee el frontend
        "bodegas": bodegas_out,
        "packs": packs_out,
    }


def _ensure_project(db: Session, proyecto_id: str) -> Proyecto:
    p = db.get(Proyecto, proyecto_id)
    if not p:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Proyecto no encontrado")
    return p


def _touch_stock(db: Session, proyecto_id: str) -> None:
    """Marca el proyecto como 'stock actualizado ahora'. Solo se llama desde las
    mutaciones de unidades (alta/edición/borrado), para que stock_updated_at refleje
    SOLO cambios de stock. Bulk update (no dispara onupdate) → setea ambos a mano."""
    now = datetime.utcnow()
    db.query(Proyecto).filter(Proyecto.id == proyecto_id).update(
        {Proyecto.stock_updated_at: now, Proyecto.updated_at: now},
        synchronize_session=False,
    )


def _num_or_none(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


@router.get("", response_model=List[UnidadOut])
def listar(proyecto_id: str, db: Session = Depends(get_db), _: Usuario = Depends(stock_access)):
    _ensure_project(db, proyecto_id)
    return db.query(Unidad).filter(Unidad.proyecto_id == proyecto_id).order_by(Unidad.numero).all()


@router.post("", response_model=UnidadOut, status_code=status.HTTP_201_CREATED)
def crear(
    proyecto_id: str,
    body: UnidadIn,
    db: Session = Depends(get_db),
    _: Usuario = Depends(stock_access),
):
    _ensure_project(db, proyecto_id)
    if db.query(Unidad).filter(Unidad.proyecto_id == proyecto_id, Unidad.numero == body.numero).first():
        raise HTTPException(status.HTTP_409_CONFLICT, detail=f"Ya existe la unidad {body.numero}")
    u = Unidad(id="u-" + uuid.uuid4().hex[:10], proyecto_id=proyecto_id, **body.model_dump())
    db.add(u)
    _touch_stock(db, proyecto_id)
    db.commit()
    db.refresh(u)
    return u


@router.put("/{unidad_id}", response_model=UnidadOut)
def actualizar(
    proyecto_id: str,
    unidad_id: str,
    body: UnidadIn,
    db: Session = Depends(get_db),
    _: Usuario = Depends(stock_access),
):
    u = db.get(Unidad, unidad_id)
    if not u or u.proyecto_id != proyecto_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Unidad no encontrada")
    for k, v in body.model_dump().items():
        setattr(u, k, v)
    _touch_stock(db, proyecto_id)
    db.commit()
    db.refresh(u)
    return u


@router.delete("/{unidad_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar(
    proyecto_id: str,
    unidad_id: str,
    db: Session = Depends(get_db),
    _: Usuario = Depends(stock_access),
):
    u = db.get(Unidad, unidad_id)
    if not u or u.proyecto_id != proyecto_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Unidad no encontrada")
    db.delete(u)
    _touch_stock(db, proyecto_id)
    db.commit()


# ── Excel formato JetBrokers v2.4 (4 sheets) ────────────────────────────────
# Estructura JB:
#   INSTRUCCIONES: descripción + guía
#   UNIDAD:        row1=marcadores REQ/OPC, row2=labels español, row3+=data deptos
#   ESTACIONAMIENTOS / BODEGAS: row1=labels español, row2+=data
# Generamos compatible con JB para que el archivo pueda subirse tanto a bc-api
# como al editor JetBrokers original.

# Labels JB para sheet UNIDAD (orden = columnas). bc-api key + label visible + REQ.
JB_UNIDAD_COLS = [
    ("numero_depto",    "Unidad Número",          "REQ"),
    ("_dormitorios",    "Dormitorios",            "REQ"),
    ("_banos",          "Baños",                  "REQ"),
    ("modelo",          "Modelo",                 "OPC"),
    ("orientacion",     "Orientacion",            "OPC"),
    ("sup_interior",    "Sup Interior",           "OPC"),
    ("sup_terraza",     "Sup Terraza",            "OPC"),
    ("sup_logia",       "Sup Logia",              "OPC"),
    ("sup_jardin",      "Sup Jardin",             "OPC"),
    ("sup_total",       "Sup Total",              "REQ"),
    ("precio_lista_uf", "ValorUF",                "REQ"),
    ("descuento_pct",   "Descuento",              "OPC"),
    ("bono_pie_pct",    "Bonopie",                "OPC"),
    ("estac_flag",      "Cotiza Estacionamiento", "OPC"),
    ("bodega_flag",     "Cotiza Bodega",          "OPC"),
    ("pack_flag",       "Cotiza Pack",            "OPC"),
    ("disponible",      "Disponible",             "OPC"),
]

JB_FLAG_TO_LABEL = {"required": "Obligatorio", "optional": "Opcional", "never": "Nunca"}

JB_BODEGAS_COLS = [("numero", "Número"), ("precio_uf", "ValorUF"), ("superficie", "Sup Total"), ("disponible", "Disponible")]
JB_ESTAC_COLS = [("numero", "Número"), ("precio_uf", "ValorUF"), ("nivel", "Nivel"), ("tipo", "Tipo"), ("disponible", "Disponible")]


def _is_depto(u: Unidad) -> bool:
    t = (u.tipo or "").lower()
    return t in ("depto", "departamento", "apartment", "")


def _valor_cambio(old, new) -> bool:
    """True si el valor cambió realmente (normaliza None/''/float/bool)."""
    def norm(x):
        if x is None or x == "":
            return None
        if isinstance(x, bool):
            return x
        try:
            return round(float(x), 2)
        except (TypeError, ValueError):
            return str(x).strip()
    return norm(old) != norm(new)


# Etiquetas legibles de campos para el comentario del timeline
_CAMPO_LABEL = {
    "precio_lista_uf": "precio", "precio_final_uf": "precio final",
    "descuento_pct": "descuento", "bono_pie_pct": "bono pie",
    "modelo": "modelo", "tipologia": "tipología", "tipo": "tipo",
    "orientacion": "orientación",
    "sup_total": "sup. total", "sup_interior": "sup. interior",
    "sup_terraza": "sup. terraza", "sup_logia": "sup. logia", "sup_jardin": "sup. jardín",
    "disponible": "disponibilidad",
    "estac_flag": "cotiza estac", "bodega_flag": "cotiza bodega", "pack_flag": "cotiza pack",
}


def _fmt_val(campo: str, val) -> str:
    """Formatea un valor para el comentario (precios con miles, disponible legible)."""
    if val is None or val == "":
        return "—"
    if campo in ("precio_lista_uf", "precio_final_uf"):
        try:
            return f"{float(val):,.0f} UF".replace(",", ".")
        except (TypeError, ValueError):
            return str(val)
    if campo in ("descuento_pct", "bono_pie_pct"):
        try:
            return f"{float(val):g}%"
        except (TypeError, ValueError):
            return str(val)
    if campo == "disponible":
        return "disponible" if val else "no disponible"
    try:
        return f"{float(val):g}"
    except (TypeError, ValueError):
        return str(val)


def _desc_modificacion(num: str, campos: list) -> str:
    """'52A: precio 3.030→3.050 UF, descuento 0%→5%'."""
    partes = [
        f"{_CAMPO_LABEL.get(k, k)} {_fmt_val(k, old)}→{_fmt_val(k, new)}"
        for k, old, new in campos
    ]
    return f"{num}: " + ", ".join(partes)


def _parse_dorm_banos(tipologia: str) -> tuple[int | None, int | None]:
    """'3D - 2B' / '3D2B' → (3, 2)."""
    import re as _re
    if not tipologia: return (None, None)
    m = _re.search(r"(\d+)\s*D[^\d]*(\d+)\s*B", tipologia.upper())
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


@router.get("/excel/template")
def descargar_plantilla(
    proyecto_id: str,
    con_datos: bool = False,
    formato: str = "jb",  # "jb" (4 sheets compat JetBrokers) | "legacy" (1 sheet vieja)
    db: Session = Depends(get_db),
    _: Usuario = Depends(stock_access),
):
    """Excel en formato JetBrokers (4 sheets: INSTRUCCIONES / UNIDAD / ESTACIONAMIENTOS / BODEGAS).

    Compatible con el editor JetBrokers — se puede descargar de bc-api,
    editar offline, y re-subir a JB o a bc-api.

    Parámetros:
      - con_datos: True → exporta unidades/bodegas/estac existentes
                   False → plantilla vacía con 1 fila de ejemplo
      - formato:   "jb" (default, 4 sheets) | "legacy" (1 sheet técnica)
    """
    proy = _ensure_project(db, proyecto_id)

    if formato == "legacy":
        return _excel_legacy(proy, proyecto_id, con_datos, db)

    wb = Workbook()
    wb.remove(wb.active)  # quitar Sheet default

    # ── 1. INSTRUCCIONES ──────────────────────────────────────────────
    ws = wb.create_sheet("INSTRUCCIONES")
    ws.append(["Stock — formato compatible JetBrokers"])
    ws.append([])
    ws.append([f"Proyecto: {proy.nombre or '?'}"])
    ws.append([f"Inmobiliaria: {proy.inmobiliaria or '?'}"])
    ws.append([f"Comuna: {proy.comuna or '?'}"])
    ws.append([])
    ws.append(["Pestañas:"])
    ws.append(["  UNIDAD            → departamentos (1 fila por depto)"])
    ws.append(["  ESTACIONAMIENTOS  → estacionamientos"])
    ws.append(["  BODEGAS           → bodegas"])
    ws.append([])
    ws.append(["Convenciones:"])
    ws.append(["  REQ = obligatorio · OPC = opcional (marcado en fila 1 de UNIDAD)"])
    ws.append(["  Cotiza Estac/Bodega/Pack: Obligatorio · Opcional · Nunca"])
    ws.append(["  Disponible: TRUE = a la venta · FALSE = reservada/vendida"])
    ws.append([])
    ws.append(["Generado por BigCapital · bc-api"])

    # ── 2. UNIDAD ─────────────────────────────────────────────────────
    ws = wb.create_sheet("UNIDAD")
    # Row 1: marcadores REQ/OPC (como JB)
    ws.append([req for (_, _, req) in JB_UNIDAD_COLS])
    # Row 2: labels en español (como JB)
    ws.append([label for (_, label, _) in JB_UNIDAD_COLS])

    if con_datos:
        unidades = db.query(Unidad).filter(Unidad.proyecto_id == proyecto_id).order_by(Unidad.numero).all()
        for u in unidades:
            if not _is_depto(u):
                continue
            dorm, banos = _parse_dorm_banos(u.tipologia or "")
            row = []
            for key, _label, _req in JB_UNIDAD_COLS:
                if key == "_dormitorios": row.append(dorm)
                elif key == "_banos": row.append(banos)
                elif key == "numero_depto": row.append(u.numero)
                elif key in ("estac_flag", "bodega_flag", "pack_flag"):
                    row.append(JB_FLAG_TO_LABEL.get(getattr(u, key) or "optional", "Opcional"))
                elif key == "disponible":
                    row.append("TRUE" if u.disponible else "FALSE")
                else:
                    row.append(getattr(u, key, None))
            ws.append(row)
    else:
        # Fila de ejemplo
        ws.append(["101", 3, 2, "A1", "N", 60.0, 2.5, 0, 0, 62.5, 3200, 0, 20,
                   "Opcional", "Obligatorio", "Nunca", "TRUE"])

    # ── 3. ESTACIONAMIENTOS ───────────────────────────────────────────
    ws = wb.create_sheet("ESTACIONAMIENTOS")
    # JB tiene row 1 vacío opcional. Usamos solo row 2 con labels.
    ws.append([])  # row 1: vacía (compat JB)
    ws.append([label for (_, label) in JB_ESTAC_COLS])
    if con_datos:
        extra = proy.extra or {}
        for source in (extra.get("estacionamientos_dom"), extra.get("_estacionamientos_dom")):
            if not source: continue
            for e in source:
                cells = e.get("cells") if isinstance(e, dict) else None
                if cells:
                    # Cells JB típicas: [_, numero, precio, nivel, tipo, ...]
                    ws.append([cells[1] if len(cells)>1 else "",
                               cells[2] if len(cells)>2 else "",
                               cells[3] if len(cells)>3 else "",
                               cells[4] if len(cells)>4 else "",
                               "TRUE" if e.get("disponible", True) else "FALSE"])
            break
        # También unidades con tipo="Estacionamiento"
        for u in db.query(Unidad).filter(Unidad.proyecto_id == proyecto_id).all():
            t = (u.tipo or "").lower()
            if "estac" in t or "parking" in t:
                ws.append([u.numero, u.precio_lista_uf, u.orientacion or "", u.modelo or "",
                           "TRUE" if u.disponible else "FALSE"])
    else:
        ws.append(["E-101", 250, "S1", "Doble", "TRUE"])

    # ── 4. BODEGAS ────────────────────────────────────────────────────
    ws = wb.create_sheet("BODEGAS")
    ws.append([])  # row 1: vacía
    ws.append([label for (_, label) in JB_BODEGAS_COLS])
    if con_datos:
        extra = proy.extra or {}
        for source in (extra.get("bodegas_dom"), extra.get("_bodegas_dom")):
            if not source: continue
            for b in source:
                cells = b.get("cells") if isinstance(b, dict) else None
                if cells:
                    ws.append([cells[1] if len(cells)>1 else "",
                               cells[2] if len(cells)>2 else "",
                               cells[3] if len(cells)>3 else "",
                               "TRUE" if b.get("disponible", True) else "FALSE"])
            break
        for u in db.query(Unidad).filter(Unidad.proyecto_id == proyecto_id).all():
            t = (u.tipo or "").lower()
            if "bodega" in t or t == "storage":
                ws.append([u.numero, u.precio_lista_uf, u.sup_total or "",
                           "TRUE" if u.disponible else "FALSE"])
    else:
        ws.append(["B-15", 80, 3.5, "TRUE"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = "_con_datos" if con_datos else ""
    safe_name = "".join(c if c.isalnum() else "_" for c in (proy.nombre or "proyecto"))
    fname = f"{safe_name}_stock{suffix}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _excel_legacy(proy, proyecto_id, con_datos, db):
    """Formato viejo de 1 sheet — fallback opcional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(HEADERS)
    if con_datos:
        for u in db.query(Unidad).filter(Unidad.proyecto_id == proyecto_id).order_by(Unidad.numero).all():
            ws.append([
                u.numero, u.modelo, u.tipologia, u.tipo, u.orientacion,
                u.sup_total, u.sup_interior, u.sup_terraza, u.sup_logia, u.sup_jardin,
                u.precio_lista_uf, u.descuento_pct, u.bono_pie_pct, u.precio_final_uf,
                u.estac_flag, u.bodega_flag, u.pack_flag,
                "TRUE" if u.disponible else "FALSE",
            ])
    else:
        ws.append(["101", "A1", "3D - 2B", "Depto", "N", 62.5, 60.0, 2.5, 0, 0,
                   3200, 0, 20, 3200, "optional", "never", "never", "TRUE"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = "_con_datos" if con_datos else ""
    safe_name = "".join(c if c.isalnum() else "_" for c in (proy.nombre or "proyecto"))
    fname = f"{safe_name}_stock_legacy{suffix}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/excel/upload")
async def subir_excel(
    proyecto_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(stock_access),
):
    """Upsert por numero_depto. Devuelve resumen {inserted, updated, errors}.

    - Preserva campos manuales (orientación) cuando el Excel viene vacío.
    - Da de baja (disponible=False) los deptos que ya no vienen en el Excel.
    - Registra un evento en extra.timeline ("Excel Stock") con qué cambió.
    """
    proy = _ensure_project(db, proyecto_id)
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="El archivo debe ser .xlsx o .xls")

    body = await file.read()
    try:
        wb = load_workbook(io.BytesIO(body), data_only=True, read_only=False)
    except Exception as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=f"Excel inválido: {e}")

    # ── Detectar formato: JB (4 sheets) o bc-api (1 sheet con headers en row 1) ──
    is_jb = _is_jb_excel(wb)
    inserted, updated, errors = [], [], []
    modificadas = []  # deptos existentes cuyos datos REALMENTE cambiaron (precio, etc.)
    nuevos_info = {}  # {numero: (modelo, precio)} de los deptos insertados, para la referencia
    deptos_con_warning: list[dict] = []  # ext. Maestra: [{numero, modelo}, ...] no registrados en extra.modelos
    by_num = {u.numero: u for u in db.query(Unidad).filter(Unidad.proyecto_id == proyecto_id).all()}
    jb_extras: dict = {}

    if is_jb:
        # Parser JB
        unidad_rows, parse_errors = _parse_jb_excel(wb)
        errors.extend(parse_errors)
        if not unidad_rows:
            # Diagnóstico: capturar sheet names + labels detectadas + sample row para debug
            try:
                ws_dbg = wb["UNIDAD"]
                rows_dbg = list(ws_dbg.iter_rows(values_only=True))
                sheets = list(wb.sheetnames)
                nrows = len(rows_dbg)
                row0 = [str(c)[:30] if c is not None else "" for c in (rows_dbg[0] if rows_dbg else [])][:20]
                row1 = [str(c)[:30] if c is not None else "" for c in (rows_dbg[1] if len(rows_dbg) > 1 else [])][:20]
                row2 = [str(c)[:30] if c is not None else "" for c in (rows_dbg[2] if len(rows_dbg) > 2 else [])][:20]
                row3 = [str(c)[:30] if c is not None else "" for c in (rows_dbg[3] if len(rows_dbg) > 3 else [])][:20]
                detail = (f"Sheet UNIDAD vacío o sin filas válidas. "
                          f"sheets={sheets} nrows={nrows} row0={row0} row1={row1} row2={row2} row3={row3} "
                          f"parse_errors={parse_errors[:3]}")
            except Exception as _e:
                detail = f"Sheet UNIDAD vacío o sin filas válidas (err diag: {_e})"
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=detail)
        # Estacionamientos + Bodegas + Packs → extra (todavía no entidades separadas)
        jb_estac = _parse_jb_estacionamientos(wb)
        jb_bodegas = _parse_jb_bodegas(wb)
        jb_packs = _parse_jb_packs(wb)  # opcional (ext. Maestra)
        # Construir shape final que el frontend lee (extra.estacionamientos / .bodegas / .packs).
        # Si hay hoja PACKS, se usa como fuente autoritativa; si no, se cae al fallback
        # de agrupar por Pack Número de las hojas individuales.
        jb_extras = _build_jb_extras(jb_estac, jb_bodegas, jb_packs)
        rows_iter = [(i + 3, r) for i, r in enumerate(unidad_rows)]
    else:
        # Parser legado bc-api
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel vacío")
        header = [str(c or "").strip() for c in rows[0]]
        if "numero_depto" not in header:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                detail="Falta columna 'numero_depto'. Descarga la plantilla para ver el formato esperado.",
            )
        idx = {h: i for i, h in enumerate(header)}
        legacy_rows = []
        for i, r in enumerate(rows[1:], start=2):
            num = str(r[idx["numero_depto"]] or "").strip()
            if not num:
                errors.append(f"Fila {i}: falta número")
                continue

            def col(name, transform=lambda v: v):
                j = idx.get(name)
                return transform(r[j]) if j is not None and j < len(r) else None

            # Regla del usuario: aparece en Excel = disponible. Ignoramos la
            # columna 'Disponible' (los humanos a veces escriben 'no disp' en
            # una fila que YA NO debería estar — eso es error del Excel).
            disp = True
            d = dict(
                numero_depto=num,
                modelo=col("modelo") or "",
                tipologia=col("tipologia") or "",
                tipo=col("tipo") or "Depto",
                orientacion=col("orientacion") or "",
                sup_total=col("sup_total"),
                sup_interior=col("sup_interior"),
                sup_terraza=col("sup_terraza"),
                sup_logia=col("sup_logia"),
                sup_jardin=col("sup_jardin"),
                precio_lista_uf=col("precio_lista_uf"),
                descuento_pct=col("descuento_pct"),
                bono_pie_pct=col("bono_pie_pct"),
                precio_final_uf=col("precio_final_uf"),
                estac_flag=col("estac_flag") or "optional",
                bodega_flag=col("bodega_flag") or "optional",
                pack_flag=col("pack_flag") or "optional",
                disponible=disp,
            )
            legacy_rows.append((i, d))
        rows_iter = legacy_rows

    # ── Auto-asignación de modelo por tipología ──
    # Excel JB no exporta columna "Modelo" (o viene vacía). Antes de procesar las filas,
    # construimos un índice {(dormitorios, banos) → nombre del modelo} a partir de
    # extra.modelos del proyecto, para autoasignar modelo a cada fila que venga sin él.
    # Si hay AMBIGÜEDAD (varios modelos con misma (D,B)) NO autoasignamos — preferimos
    # dejar modelo="" antes que mapear mal.
    _ex_modelos = (proy.extra or {}).get("modelos") or []
    _by_db: dict[tuple[int, int], str] = {}
    _ambiguos: set[tuple[int, int]] = set()
    for _m in _ex_modelos:
        if not isinstance(_m, dict):
            continue
        try:
            _d = int(_m.get("dormitorios")) if _m.get("dormitorios") is not None else None
            _b = int(_m.get("banos")) if _m.get("banos") is not None else None
        except (ValueError, TypeError):
            _d = _b = None
        _n = (_m.get("nombre") or "").strip()
        if _d is None or _b is None or not _n:
            continue
        key = (_d, _b)
        if key in _by_db and _by_db[key] != _n:
            _ambiguos.add(key)
        else:
            _by_db.setdefault(key, _n)
    # quitar ambiguos del mapa
    for k in _ambiguos:
        _by_db.pop(k, None)

    # ── Procesar filas (mismo loop para JB o legacy) ──
    auto_asignados = 0
    for i, d in rows_iter:
        num = str(d.get("numero_depto") or "").strip()
        if not num:
            errors.append(f"Fila {i}: falta 'Unidad Número'")
            continue
        # auto-asignar modelo si viene vacío y la tipología permite inferir (D,B)
        _modelo_raw = (d.get("modelo") or "").strip()
        if not _modelo_raw:
            _tip = (d.get("tipologia") or "").strip()
            _dorm, _banos = _parse_dorm_banos(_tip)
            if _dorm is not None and _banos is not None:
                _mname = _by_db.get((_dorm, _banos))
                if _mname:
                    d["modelo"] = _mname
                    auto_asignados += 1
        data = dict(
            numero=num,
            modelo=d.get("modelo") or "",
            tipologia=d.get("tipologia") or "",
            tipo=d.get("tipo") or "Depto",
            orientacion=d.get("orientacion") or "",
            sup_total=_num_or_none(d.get("sup_total")),
            sup_interior=_num_or_none(d.get("sup_interior")),
            sup_terraza=_num_or_none(d.get("sup_terraza")),
            sup_logia=_num_or_none(d.get("sup_logia")),
            sup_jardin=_num_or_none(d.get("sup_jardin")),
            precio_lista_uf=_num_or_none(d.get("precio_lista_uf")),
            descuento_pct=_num_or_none(d.get("descuento_pct")) or 0,
            bono_pie_pct=_num_or_none(d.get("bono_pie_pct")) or 0,
            precio_final_uf=_num_or_none(d.get("precio_final_uf")),
            estac_flag=d.get("estac_flag") or "optional",
            bodega_flag=d.get("bodega_flag") or "optional",
            pack_flag=d.get("pack_flag") or "optional",
            # Regla: aparece en el Excel = disponible. d.get('disponible') ya
            # viene True por _parse_jb_excel (default) salvo que JB marque FALSE
            # explícito (caso pack), en cuyo caso respetamos.
            disponible=bool(d.get("disponible", True)),
        )

        if num in by_num:
            u = by_num[num]
            # Upsert parcial: campos manuales que el origen (PlanOk/MNK) NO provee
            # se preservan si vienen vacíos, para no pisar datos cargados a mano.
            # Ej: orientación — PlanOk no la expone, es manual en BC.
            PRESERVAR_SI_VACIO = {"orientacion"}
            cambios_campos = []  # (campo, old, new) de lo que cambió de verdad
            for k, v in data.items():
                if k in PRESERVAR_SI_VACIO and (v is None or v == ""):
                    continue
                old = getattr(u, k, None)
                if _valor_cambio(old, v):
                    cambios_campos.append((k, old, v))
                setattr(u, k, v)
            updated.append(num)
            if cambios_campos:
                modificadas.append((num, cambios_campos))
        else:
            u = Unidad(id="u-" + uuid.uuid4().hex[:10], proyecto_id=proyecto_id, **data)
            db.add(u)
            inserted.append(num)
            nuevos_info[num] = (data.get("modelo") or "", data.get("precio_lista_uf"))

        # Ext. Maestra: si el Excel marcó _modelo_warning para este depto, anotarlo
        if is_jb and str(d.get("_modelo_warning") or "").strip() in ("1", "true", "True"):
            deptos_con_warning.append({"numero": num, "modelo": data.get("modelo") or ""})

    # ── Baja de stock: lo que NO viene en el Excel → disponible=False ──
    # El Excel del scraper sube el stock vigente; lo que falta = vendido/reservado.
    # Aplica a deptos (regla pedida por el usuario: 'aparece en Excel = disponible').
    # NO se borra (preserva registro + datos manuales). Si reaparece → se reactiva.
    seen_nums = {str(d.get("numero_depto") or "").strip() for _i, d in rows_iter}
    dados_de_baja = []
    for u in db.query(Unidad).filter(Unidad.proyecto_id == proyecto_id).all():
        if _is_depto(u) and u.numero not in seen_nums and u.disponible:
            u.disponible = False
            dados_de_baja.append(u.numero)

    proy.stock_last_upload = {
        "filename": file.filename,
        "at": datetime.utcnow().isoformat() + "Z",
        "inserted": len(inserted),
        "updated": len(updated),
        "errors": len(errors),
        "dados_de_baja": len(dados_de_baja),
        "format": "jb_v2.4" if is_jb else "bc_api",
    }
    proy.stock_updated_at = datetime.utcnow()  # 1.12 · marca el cambio de stock

    # ── Timeline: comentario "Sin cambios" o resumen de qué cambió ──
    _email_usuario = (getattr(usuario, "email", "") or "").lower()
    if _email_usuario.startswith("mnk-scraper"):
        _origen = "Actualización automática (scraper MNK · PlanOk)"
    elif _email_usuario.startswith("maestra-scraper"):
        _origen = "Actualización automática (Maestra · Excel)"
    else:
        _origen = "Carga de Excel de stock"

    def _corta(lst):
        return ", ".join(lst[:6]) + ("…" if len(lst) > 6 else "")

    if not (inserted or dados_de_baja or modificadas):
        # Nada cambió respecto al estado anterior
        _depto_pl = "depto" if len(updated) == 1 else "deptos"
        _detalles = f"{_origen} — Sin cambios (stock idéntico: {len(updated)} {_depto_pl})"
    else:
        _partes = []
        if inserted:
            _t = "depto nuevo" if len(inserted) == 1 else "deptos nuevos"
            _refs = []
            for n in inserted[:6]:
                _modelo, _precio = nuevos_info.get(n, ("", None))
                _ex = [x for x in (_modelo, _fmt_val("precio_lista_uf", _precio) if _precio else "") if x]
                _refs.append(f"{n} ({', '.join(_ex)})" if _ex else n)
            _det_new = ", ".join(_refs) + ("…" if len(inserted) > 6 else "")
            _partes.append(f"{len(inserted)} {_t} → {_det_new}")
        if dados_de_baja:
            _t = "dado de baja" if len(dados_de_baja) == 1 else "dados de baja"
            _partes.append(f"{len(dados_de_baja)} {_t} ({_corta(dados_de_baja)})")
        if modificadas:
            _t = "modificado" if len(modificadas) == 1 else "modificados"
            _descs = [_desc_modificacion(num, campos) for num, campos in modificadas]
            _det_mod = " · ".join(_descs[:5]) + ("…" if len(_descs) > 5 else "")
            _partes.append(f"{len(modificadas)} {_t} → {_det_mod}")
        _detalles = f"{_origen} — " + ", ".join(_partes)

    _evento = {
        "id": "tl-" + uuid.uuid4().hex[:10],
        "fecha": datetime.utcnow().isoformat() + "Z",
        "tipo": "Excel Stock",
        "detalles": _detalles,
        "usuario": getattr(usuario, "email", None) or "sistema",
        "archivo_url": None,
    }

    # Componer extra: jb_extras (si JB) + timeline actualizado
    _extra = {**(proy.extra or {})}
    if is_jb and jb_extras:
        _extra.update(jb_extras)
    # Ext. Maestra: persistir deptos con warning de modelo (el frontend muestra badge rojo).
    # Si no hay warnings, limpiamos la clave (un re-upload sano no debe dejar warnings viejos).
    if is_jb:
        if deptos_con_warning:
            _extra["_deptos_con_warning"] = deptos_con_warning
        else:
            _extra.pop("_deptos_con_warning", None)
    _tl = list(_extra.get("timeline") or [])
    _tl.insert(0, _evento)
    # Si hay warnings, agregar también un evento rojo destacado en la timeline.
    if is_jb and deptos_con_warning:
        _warn_evt = {
            "id": "tl-" + uuid.uuid4().hex[:10],
            "fecha": datetime.utcnow().isoformat() + "Z",
            "tipo": "Alerta",
            "severity": "CRITICO",
            "titulo": f"Modelos no registrados detectados: {len({d['modelo'] for d in deptos_con_warning})}",
            "detalles": (
                f"⚠ {len(deptos_con_warning)} depto(s) con modelo fuera del catálogo de "
                f"este proyecto: "
                + ", ".join(f"{d['numero']} ({d['modelo']})" for d in deptos_con_warning[:8])
                + ("…" if len(deptos_con_warning) > 8 else "")
                + ". Estos deptos quedan visibles con badge de advertencia hasta que "
                "registres los modelos en este proyecto."
            ),
            "usuario": getattr(usuario, "email", None) or "sistema",
            "archivo_url": None,
        }
        _tl.insert(0, _warn_evt)
    _extra["timeline"] = _tl
    proy.extra = _extra

    db.commit()

    background_tasks.add_task(
        email_service.notify_change, "Stock actualizado", proy.nombre or proyecto_id,
        f"Subida de Excel: +{len(inserted)} nuevas, {len(updated)} actualizadas"
        + (f", {dados_de_baja} dadas de baja" if dados_de_baja else "") + ".",
        proyecto_id,
    )
    return {
        "format": "jb_v2.4" if is_jb else "bc_api",
        "inserted": len(inserted),
        "updated": len(updated),
        "dados_de_baja": dados_de_baja,
        "errors": errors,
        "estacionamientos_count": len(jb_extras.get("_estacionamientos_dom", [])),
        "bodegas_count": len(jb_extras.get("_bodegas_dom", [])),
        "warnings_count": len(deptos_con_warning) if is_jb else 0,
        "stock_last_upload": proy.stock_last_upload,
    }


# ── Endpoint para alertas externas (scrapers) ───────────────────────────────
# El scraper de Maestra postea aquí cuando detecta algo crítico que NO está asociado
# a un upload (login falló, Excel no se encontró, etc.). El frontend lo renderiza
# como un evento rojo destacado en la timeline.

@router.post("/timeline/alerta", status_code=status.HTTP_201_CREATED)
def crear_alerta_timeline(
    proyecto_id: str,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(stock_access),
):
    """Inserta un evento de tipo 'Alerta' en extra.timeline del proyecto.

    Body esperado: {severity: "CRITICO|WARNING", titulo: str, detalle: str}
    """
    proy = _ensure_project(db, proyecto_id)
    severity = str(payload.get("severity") or "WARNING").upper()
    if severity not in ("CRITICO", "WARNING"):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"severity inválida: {severity!r} (esperaba CRITICO o WARNING)",
        )
    titulo = str(payload.get("titulo") or "").strip()
    detalle = str(payload.get("detalle") or "").strip()
    if not titulo:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Falta 'titulo'")

    evento = {
        "id": "tl-" + uuid.uuid4().hex[:10],
        "fecha": datetime.utcnow().isoformat() + "Z",
        "tipo": "Alerta",
        "severity": severity,
        "titulo": titulo,
        "detalles": detalle or titulo,
        "usuario": getattr(usuario, "email", None) or "sistema",
        "archivo_url": None,
    }
    _extra = {**(proy.extra or {})}
    _tl = list(_extra.get("timeline") or [])
    _tl.insert(0, evento)
    _extra["timeline"] = _tl
    proy.extra = _extra
    db.commit()
    return {"ok": True, "evento_id": evento["id"]}
