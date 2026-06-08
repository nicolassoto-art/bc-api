"""Inbox processor · 2026-06-08

Lee la bandeja de entrada de sistema@bigcapital.cl vía IMAP. Para cada email
nuevo desde un remitente autorizado (nicolas.soto@bigcapital.cl por default)
con un adjunto .xlsx/.xls:

1. Identifica el proyecto destino (cascada):
   - Subject "id:<jb-xxx>" → match exacto por id
   - Subject contiene nombre del proyecto (fuzzy match con allProjects)
   - Nombre del archivo (ej. "stock-pinar.xlsx") → fuzzy match
2. Aplica el Excel al proyecto (reusa POST /proyectos/{id}/unidades/excel).
3. Responde al remitente con email de confirmación (✓ aplicado · N insertadas,
   M actualizadas, X errores) o detalle del error.
4. Marca el email como leído (Seen) para no procesarlo dos veces.

Disparado por APScheduler cada inbox_poll_minutes (default 10 min) desde main.py.
También accesible vía endpoint manual POST /admin/inbox/poll (super_admin) para
testing.

SEGURIDAD:
- Solo procesa emails desde inbox_allowed_senders (default nicolas.soto).
- El service_token de bc-api es interno y no se expone al exterior.
- IMAP con TLS (puerto 993) + App Password de Gmail (sin OAuth interactivo).
"""
from __future__ import annotations
import email
import email.utils
import imaplib
import io
import logging
import re
import smtplib
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from email.utils import formataddr
from html import escape
from typing import Optional

import httpx
from sqlalchemy.orm import Session

from app.db import SessionLocal
from app.models.proyecto import Proyecto
from app.services.email_service import _configured, _fecha_cl
from app.settings import settings

log = logging.getLogger(__name__)


@dataclass
class InboxEmail:
    uid: bytes
    from_addr: str
    from_name: str
    subject: str
    body: str  # texto plano del cuerpo (incluye Fwd: con info del remitente original)
    date: Optional[datetime]
    attachments: list  # [(filename, bytes)]


# ── Conexión IMAP ────────────────────────────────────────────────────────


def _imap_creds() -> tuple[str, str, str, int]:
    """Devuelve (user, pass, host, port). Usa imap_* si están seteados, sino smtp_*."""
    user = (settings.imap_user or settings.smtp_user).strip()
    pwd = (settings.imap_pass or settings.smtp_pass).replace(" ", "")  # Gmail App Pwd sin espacios
    host = settings.imap_host or "imap.gmail.com"
    port = settings.imap_port or 993
    return user, pwd, host, port


def _imap_configured() -> bool:
    user, pwd, _, _ = _imap_creds()
    return bool(user and pwd)


# ── Normalización + fuzzy matching ───────────────────────────────────────


def _norm(s: str) -> str:
    """Normaliza para comparar: minúsculas, sin acentos, sin puntuación."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return s.strip()


def _extract_from_original(body: str) -> Optional[str]:
    """Extrae el From original de un email reenviado. Busca patrones tipo
    'From: xxx@yyy.com' o 'De: xxx@yyy.com' que aparecen en el header del Fwd:."""
    if not body:
        return None
    # Match 'From: Nombre <email@dom>' o 'De: email@dom' (ES + EN)
    patterns = [
        r"(?:^|\n)\s*(?:From|De|FROM|DE)\s*[:>]\s*(?:[^<\n]*<)?([\w._%+-]+@[\w.-]+\.[a-zA-Z]{2,})",
        r"<\s*([\w._%+-]+@[\w.-]+\.[a-zA-Z]{2,})\s*>",
    ]
    for pat in patterns:
        m = re.search(pat, body)
        if m:
            return m.group(1).strip().lower()
    return None


def _dominio_de(email_addr: str) -> Optional[str]:
    """Extrae el dominio de un email (ej. 'juan@vallatrix.cl' → 'vallatrix.cl').
    Filtra dominios genéricos (@gmail.com, @hotmail.com, etc.) porque no identifican."""
    if not email_addr or "@" not in email_addr:
        return None
    dom = email_addr.split("@", 1)[1].strip().lower()
    # Filtrar dominios genéricos personales
    GENERICOS = {
        "gmail.com", "googlemail.com", "hotmail.com", "outlook.com",
        "yahoo.com", "yahoo.es", "live.com", "icloud.com", "me.com",
        "bigcapital.cl",  # del propio Nicolás
    }
    if dom in GENERICOS:
        return None
    return dom


def _extraer_nombre_proyecto_del_excel(body: bytes) -> str:
    """Abre el Excel y devuelve un string concatenado con TODO el contenido de las
    primeras ~15 filas × 10 cols de cada hoja. Útil para hacer fuzzy match con
    nombre de proyecto/inmobiliaria. Devuelve '' si el archivo es inválido."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(body), data_only=True, read_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                parts.append(sheet_name)
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx >= 15:  # primeras 15 filas suelen tener header/título
                        break
                    for cell in row[:10]:
                        if cell and isinstance(cell, str) and len(cell) >= 3:
                            parts.append(cell)
            except Exception:
                continue
        return " ".join(parts)
    except Exception as e:
        log.info("inbox: no se pudo leer contenido del Excel para match: %s", e)
        return ""


def _extract_body_text(msg) -> str:
    """Extrae texto plano del body del email (incluye el contenido del Fwd: original)."""
    if not msg:
        return ""
    out = []
    for part in msg.walk():
        ctype = part.get_content_type()
        if ctype in ("text/plain", "text/html") and part.get_content_disposition() != "attachment":
            try:
                payload = part.get_payload(decode=True) or b""
                charset = part.get_content_charset() or "utf-8"
                txt = payload.decode(charset, errors="replace")
                # Strip HTML rudimentario si vino en text/html
                if ctype == "text/html":
                    txt = re.sub(r"<[^>]+>", " ", txt)
                out.append(txt)
            except Exception:
                continue
    return " ".join(out)


def _match_proyecto(
    db: Session,
    subject: str,
    filename: str,
    body: str = "",
    excel_bytes: bytes = b"",
) -> tuple[Optional[Proyecto], str]:
    """Identifica el proyecto destino con cascada de heurísticas.

    Las inmobiliarias generalmente mandan sin el nombre del proyecto en el subject
    ('Stock actualizado', 'Stock semana 23'). Cuando Nicolás reenvía, el cuerpo
    del email contiene el Fwd con info del remitente original (inmobiliaria).
    Por eso buscamos en SUBJECT + FILENAME + BODY, y como último recurso usamos
    la inmobiliaria (si tiene 1 solo proyecto activo, match único).

    Devuelve (proyecto, motivo). El motivo es para logging/email de error.

    Orden:
    1. id:<xxx> explícito en subject/body
    2. Nombre del proyecto en subject/filename/body (fuzzy, single match)
    3. Inmobiliaria en subject/body → si tiene 1 solo proyecto activo, match
    4. None si nada matchea de forma inequívoca
    """
    # 1) ID explícito
    for src in (subject or "", body or ""):
        m = re.search(r"id\s*[:=]\s*([a-z0-9\-_]{4,})", src, re.IGNORECASE)
        if m:
            pid = m.group(1).strip()
            p = db.get(Proyecto, pid)
            if p:
                return p, f"id explícito en email ({pid})"

    activos = (
        db.query(Proyecto)
        .filter(Proyecto.activo == True, Proyecto.deleted_at.is_(None))  # noqa
        .all()
    )
    if not activos:
        return None, "no hay proyectos activos en bc-api"

    # Limitar el body a primeros 4000 chars (suficiente para el header del Fwd:)
    body_short = (body or "")[:4000]
    # Contenido del Excel (primeras 15 filas × 10 cols de cada hoja) — captura el
    # nombre del proyecto cuando aparece dentro del archivo mismo.
    excel_content = _extraer_nombre_proyecto_del_excel(excel_bytes) if excel_bytes else ""
    haystacks = [
        ("subject", _norm(subject)),
        ("filename", _norm(filename)),
        ("body", _norm(body_short)),
        ("excel", _norm(excel_content)),
    ]

    # 2) Match por nombre o id en subject/filename/body, exigiendo UNICIDAD
    for source, hay in haystacks:
        if not hay:
            continue
        candidatos = []
        for p in activos:
            nombre_norm = _norm(p.nombre or "")
            id_norm = _norm(p.id or "")
            if nombre_norm and len(nombre_norm) >= 4 and nombre_norm in hay:
                candidatos.append(p)
            elif id_norm and len(id_norm) >= 4 and id_norm in hay:
                candidatos.append(p)
        # Dedup
        seen = set()
        candidatos = [p for p in candidatos if not (p.id in seen or seen.add(p.id))]
        if len(candidatos) == 1:
            return candidatos[0], f"nombre del proyecto en {source}"
        if len(candidatos) > 1:
            # Ambiguo en esta fuente — seguimos al siguiente nivel, podría desambiguar
            log.info("inbox: %s ambiguo con %d candidatos %s — sigo", source, len(candidatos), [p.id for p in candidatos[:5]])

    # 3) Match por INMOBILIARIA: buscar el nombre de la inmobiliaria en subject+body+excel.
    #    Si solo tiene 1 proyecto activo, match único.
    full_hay = _norm(f"{subject} {filename} {body_short} {excel_content}")
    inmob_by_name: dict[str, list[Proyecto]] = {}
    for p in activos:
        inmob = _norm(p.inmobiliaria or "")
        if inmob and len(inmob) >= 3:
            inmob_by_name.setdefault(inmob, []).append(p)
    inmob_matches: list[Proyecto] = []
    for inmob_norm, proyectos in inmob_by_name.items():
        if inmob_norm in full_hay:
            inmob_matches.extend(proyectos)
    # Dedup
    seen = set()
    inmob_matches = [p for p in inmob_matches if not (p.id in seen or seen.add(p.id))]
    if len(inmob_matches) == 1:
        p = inmob_matches[0]
        return p, f"inmobiliaria '{p.inmobiliaria}' tiene solo este proyecto"
    if len(inmob_matches) > 1:
        log.info("inbox: inmobiliaria matchea con varios proyectos: %s", [p.id for p in inmob_matches[:5]])

    # 4) Match por DOMINIO del remitente original (extraído del Fwd:) contra
    #    p.inmobiliaria_web (sitio de la inmobiliaria, ej. https://vallatrix.cl)
    from_orig = _extract_from_original(body or "")
    dominio = _dominio_de(from_orig) if from_orig else None
    if dominio:
        candidatos_dom = []
        for p in activos:
            web = (p.extra or {}).get("inmobiliaria", {})
            if isinstance(web, dict):
                web = web.get("web", "")
            web_str = (p.inmobiliaria_web if hasattr(p, "inmobiliaria_web") else "") or (web or "")
            if not web_str:
                # Fallback: buscar en extra.inmobiliaria.web (estructura nested)
                inm = (p.extra or {}).get("inmobiliaria") or {}
                if isinstance(inm, dict):
                    web_str = inm.get("web") or ""
            if not web_str:
                continue
            web_norm = web_str.lower().replace("https://", "").replace("http://", "").replace("www.", "").rstrip("/")
            if dominio in web_norm or web_norm.split("/")[0].endswith(dominio):
                candidatos_dom.append(p)
        seen = set()
        candidatos_dom = [p for p in candidatos_dom if not (p.id in seen or seen.add(p.id))]
        if len(candidatos_dom) == 1:
            p = candidatos_dom[0]
            return p, f"dominio del remitente '{dominio}' matchea con la inmobiliaria de este proyecto"
        if len(candidatos_dom) > 1:
            log.info("inbox: dominio %s matchea con varios proyectos: %s", dominio, [p.id for p in candidatos_dom[:5]])

    return None, f"no se identificó el proyecto (subject='{subject[:40]}', filename='{filename}', dominio_origen='{dominio or '?'}')"


# ── Lectura del inbox ────────────────────────────────────────────────────


def _parse_message(msg) -> InboxEmail:
    """Convierte un email.message.Message en InboxEmail."""
    from_raw = msg.get("From", "")
    from_name, from_addr = email.utils.parseaddr(from_raw)
    subj = msg.get("Subject", "")
    # Decodificar subject si viene en =?utf-8?B?...?=
    parts = email.header.decode_header(subj)
    subj = "".join(
        (p.decode(c or "utf-8", errors="replace") if isinstance(p, bytes) else p)
        for p, c in parts
    )
    date_str = msg.get("Date", "")
    dt = None
    try:
        dt = email.utils.parsedate_to_datetime(date_str)
    except Exception:
        pass

    attachments = []
    for part in msg.walk():
        if part.get_content_disposition() == "attachment":
            fname = part.get_filename() or ""
            if fname:
                fparts = email.header.decode_header(fname)
                fname = "".join(
                    (p.decode(c or "utf-8", errors="replace") if isinstance(p, bytes) else p)
                    for p, c in fparts
                )
            payload = part.get_payload(decode=True) or b""
            if fname and payload:
                attachments.append((fname, payload))
    body_text = _extract_body_text(msg)
    return InboxEmail(
        uid=b"",  # se setea afuera
        from_addr=(from_addr or "").lower(),
        from_name=from_name or "",
        subject=subj,
        body=body_text,
        date=dt,
        attachments=attachments,
    )


def _fetch_new(allowed_senders: list[str]) -> list[InboxEmail]:
    """Conecta IMAP, busca UNSEEN, filtra por remitente, devuelve la lista."""
    if not _imap_configured():
        log.info("inbox: IMAP no configurado — skip")
        return []
    user, pwd, host, port = _imap_creds()
    out: list[InboxEmail] = []
    try:
        M = imaplib.IMAP4_SSL(host, port, timeout=30)
        try:
            M.login(user, pwd)
        except Exception as e:
            log.error("inbox: login IMAP falló: %s", e)
            return []
        try:
            M.select("INBOX")
            # Solo UNSEEN para no reprocesar (Gmail los marca SEEN al hacer fetch sin BODY.PEEK)
            typ, data = M.search(None, "UNSEEN")
            if typ != "OK":
                log.warning("inbox: search no-OK: %s", typ)
                return []
            uids = data[0].split() if data and data[0] else []
            log.info("inbox: %d emails UNSEEN", len(uids))
            for uid in uids:
                try:
                    # PEEK para no marcar como visto hasta que aplique exitosamente
                    typ, msg_data = M.fetch(uid, "(BODY.PEEK[])")
                    if typ != "OK" or not msg_data or not msg_data[0]:
                        continue
                    raw = msg_data[0][1]
                    if not isinstance(raw, (bytes, bytearray)):
                        continue
                    msg = email.message_from_bytes(raw)
                    item = _parse_message(msg)
                    item.uid = uid
                    if item.from_addr not in allowed_senders:
                        log.info("inbox: %s remitente NO autorizado, skip", item.from_addr)
                        continue
                    if not item.attachments:
                        log.info("inbox: %s sin adjuntos, skip", item.from_addr)
                        continue
                    out.append(item)
                except Exception as e:
                    log.warning("inbox: error parseando uid=%s: %s", uid, e)
        finally:
            try: M.close()
            except Exception: pass
            try: M.logout()
            except Exception: pass
    except Exception as e:
        log.error("inbox: error general: %s", e, exc_info=True)
    return out


def _mark_seen(uid: bytes):
    """Marca un mensaje como leído (Seen) para no reprocesar."""
    if not _imap_configured() or not uid:
        return
    user, pwd, host, port = _imap_creds()
    try:
        M = imaplib.IMAP4_SSL(host, port, timeout=15)
        M.login(user, pwd)
        M.select("INBOX")
        M.store(uid, "+FLAGS", "(\\Seen)")
        M.close()
        M.logout()
    except Exception as e:
        log.warning("inbox: no se pudo marcar seen uid=%s: %s", uid, e)


# ── Aplicación del Excel ─────────────────────────────────────────────────


def _aplicar_excel(proyecto_id: str, filename: str, body: bytes) -> dict:
    """Llama al endpoint POST /proyectos/{id}/unidades/excel internamente.

    Genera un JWT temporal del admin (settings.notify_to) y lo usa como bearer.
    Retorna el JSON de la respuesta o {error: str} si falla.
    """
    try:
        from app.services.auth import create_token
        token, _ = create_token(sub=settings.notify_to.strip().lower(), extra={"stock_admin": True})
    except Exception as e:
        return {"error": f"no se pudo crear token interno: {e}"}

    # bc-api escucha en 127.0.0.1:8001 (uvicorn) — la URL interna depende del entorno.
    # Si BC_API_INTERNAL_URL no está seteado, asumimos localhost.
    url = (getattr(settings, "bc_api_internal_url", "") or "http://127.0.0.1:8001").rstrip("/")
    endpoint = f"{url}/proyectos/{proyecto_id}/unidades/excel"
    try:
        with httpx.Client(timeout=60) as client:
            r = client.post(
                endpoint,
                headers={"Authorization": f"Bearer {token}"},
                files={"file": (filename, body, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        if r.status_code >= 400:
            try: detail = r.json().get("detail") or r.text
            except Exception: detail = r.text
            return {"error": f"HTTP {r.status_code}: {str(detail)[:200]}"}
        return r.json()
    except Exception as e:
        return {"error": f"request falló: {e}"}


# ── Email de respuesta ───────────────────────────────────────────────────


def _send_response(to_addr: str, subject: str, body_text: str, body_html: str):
    """Envía un email de respuesta al remitente original."""
    if not _configured():
        log.warning("inbox response: SMTP no configurado, no se envía.")
        return
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        from_addr = settings.smtp_from or settings.smtp_user
        msg["From"] = formataddr((settings.smtp_from_name, from_addr))
        msg["To"] = to_addr
        msg["Reply-To"] = from_addr
        msg.set_content(body_text)
        msg.add_alternative(body_html, subtype="html")
        with smtplib.SMTP(settings.smtp_host, settings.smtp_port, timeout=20) as s:
            s.starttls()
            s.login(settings.smtp_user, settings.smtp_pass.replace(" ", ""))
            s.send_message(msg)
        log.info("inbox response → %s · %s", to_addr, subject[:60])
    except Exception as e:
        log.warning("inbox response falló (%s): %s", to_addr, e)


def _html_ok(proyecto: Proyecto, filename: str, result: dict) -> str:
    inserted = len(result.get("inserted") or result.get("insertadas") or [])
    updated = len(result.get("updated") or result.get("actualizadas") or result.get("modificadas") or [])
    errors = result.get("errors") or result.get("errores") or []
    n_err = len(errors) if isinstance(errors, list) else 0
    err_html = ""
    if n_err:
        sample = errors[:5] if isinstance(errors, list) else []
        err_html = (
            f'<div style="margin-top:14px;padding:10px 14px;background:#fef3c7;border-left:3px solid #ca8a04;border-radius:6px;color:#854d0e">'
            f'<b>⚠ {n_err} errores de validación (no bloquearon el upload)</b>'
            f'<ul style="margin:6px 0 0;padding-left:18px;font-size:12.5px">'
            f'{"".join(f"<li>{escape(str(e))}</li>" for e in sample)}'
            f'{f"<li>… y {n_err-5} más</li>" if n_err > 5 else ""}'
            f'</ul></div>'
        )
    return f"""<!doctype html><html><body style="margin:0;background:#f3f4f6;padding:24px 16px;font-family:-apple-system,sans-serif">
      <div style="max-width:560px;margin:0 auto;background:#fff;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden">
        <div style="background:linear-gradient(135deg,#7DC242,#5fa832);color:#0a0d12;padding:14px 18px;font-weight:800;font-size:15px">
          ✅ Stock actualizado · {escape(proyecto.nombre or proyecto.id)}
        </div>
        <div style="padding:18px">
          <div style="margin-bottom:10px;color:#374151;font-size:13.5px">Tu archivo <b>{escape(filename)}</b> se aplicó correctamente.</div>
          <table style="width:100%;border-collapse:collapse">
            <tr><td style="padding:6px 0;color:#6b7280">Insertadas</td><td style="padding:6px 0;text-align:right;color:#16a34a;font-weight:700">{inserted}</td></tr>
            <tr><td style="padding:6px 0;color:#6b7280">Actualizadas</td><td style="padding:6px 0;text-align:right;color:#0a0d12;font-weight:700">{updated}</td></tr>
            <tr><td style="padding:6px 0;color:#6b7280">Errores</td><td style="padding:6px 0;text-align:right;color:{'#dc2626' if n_err else '#16a34a'};font-weight:700">{n_err}</td></tr>
          </table>
          {err_html}
          <p style="margin:14px 0 0;font-size:11px;color:#9ca3af">{escape(_fecha_cl())} · <a href="https://herramientas.bigcapital.cl/src/stock-interno/proyecto-vista.html?id={escape(proyecto.id)}" style="color:#1f7a3d">Ver proyecto</a></p>
        </div>
      </div></body></html>"""


def _html_err(filename: str, motivo: str, sugerencia: str = "") -> str:
    return f"""<!doctype html><html><body style="margin:0;background:#fef2f2;padding:24px 16px;font-family:-apple-system,sans-serif">
      <div style="max-width:560px;margin:0 auto;background:#fff;border:1px solid #fca5a5;border-radius:10px;overflow:hidden">
        <div style="background:#dc2626;color:#fff;padding:14px 18px;font-weight:800;font-size:15px">
          🚨 No se pudo aplicar el Excel
        </div>
        <div style="padding:18px;color:#374151;font-size:13.5px">
          <div><b>Archivo:</b> {escape(filename)}</div>
          <div style="margin-top:10px;padding:10px 14px;background:#fee2e2;border-left:3px solid #dc2626;border-radius:6px;color:#7f1d1d;white-space:pre-wrap">{escape(motivo)}</div>
          {f'<div style="margin-top:14px;color:#374151"><b>Sugerencia:</b> {escape(sugerencia)}</div>' if sugerencia else ''}
          <p style="margin:14px 0 0;font-size:11px;color:#9ca3af">{escape(_fecha_cl())}</p>
        </div>
      </div></body></html>"""


# ── Procesador principal ─────────────────────────────────────────────────


def process_inbox() -> dict:
    """Disparado por APScheduler (cada N min) o manualmente. Devuelve un resumen."""
    if not settings.inbox_processor_enabled:
        return {"enabled": False, "processed": 0}
    allowed = [e.strip().lower() for e in (settings.inbox_allowed_senders or "").split(",") if e.strip()]
    if not allowed:
        log.warning("inbox: sin remitentes autorizados — abort")
        return {"error": "sin allowed senders", "processed": 0}

    items = _fetch_new(allowed)
    log.info("inbox: %d email(es) elegibles", len(items))
    procesados = []
    for item in items:
        # Tomar el PRIMER Excel del email (si hay varios, los siguientes se ignoran).
        xlsx = next(((fn, body) for fn, body in item.attachments if fn.lower().endswith((".xlsx", ".xls"))), None)
        if not xlsx:
            log.info("inbox: %s sin Excel, skip", item.from_addr)
            continue
        filename, body = xlsx
        with SessionLocal() as db:
            proy, motivo = _match_proyecto(db, item.subject, filename, item.body, excel_bytes=body)
            if not proy:
                log.warning(
                    "inbox: no pude identificar proyecto · subject=%r filename=%r motivo=%s",
                    item.subject[:80], filename, motivo,
                )
                # Listar inmobiliarias detectadas en el email (si las hay) para que el usuario sepa
                _send_response(
                    item.from_addr,
                    f"⚠ No pude identificar el proyecto · {filename}",
                    f"No pude identificar a qué proyecto pertenece el Excel.\n"
                    f"Motivo: {motivo}\n"
                    f"Subject: {item.subject}\n"
                    f"Archivo: {filename}\n\n"
                    f"Sugerencias:\n"
                    f"• Reenvía agregando el NOMBRE del proyecto al subject (ej. 'Stock Portal del Pinar')\n"
                    f"• O el id (ej. 'id:jb-tvfylemz')\n"
                    f"• O cualquier mención del nombre del proyecto/inmobiliaria en el cuerpo del email basta\n",
                    _html_err(
                        filename,
                        f"No pude identificar el proyecto.\nMotivo: {motivo}\nSubject recibido: '{item.subject}'",
                        "Reenvía agregando el nombre del proyecto al subject (ej. 'Stock Portal del Pinar') o el id (ej. 'id:jb-tvfylemz'). También vale si mencionás el nombre del proyecto o de la inmobiliaria en el cuerpo del email.",
                    ),
                )
                _mark_seen(item.uid)
                procesados.append({"from": item.from_addr, "result": "no_proyecto"})
                continue
            # Aplicar el Excel
            res = _aplicar_excel(proy.id, filename, body)
            if "error" in res:
                log.warning("inbox: aplicar Excel falló · proy=%s · %s", proy.id, res["error"])
                _send_response(
                    item.from_addr,
                    f"🚨 Error aplicando Excel · {proy.nombre or proy.id}",
                    f"No se pudo aplicar el Excel '{filename}' al proyecto {proy.nombre or proy.id}.\n\nError: {res['error']}",
                    _html_err(filename, res["error"], "Verifica el formato del Excel (debe ser plantilla JB v2.4 o el formato bc-api). Recarga el editor y prueba subirlo manualmente para ver el detalle del error."),
                )
                _mark_seen(item.uid)
                procesados.append({"from": item.from_addr, "result": "error", "proy": proy.id, "msg": res["error"]})
                continue
            # OK
            log.info("inbox: ✓ aplicado %s · proy=%s", filename, proy.id)
            _send_response(
                item.from_addr,
                f"✅ Stock actualizado · {proy.nombre or proy.id}",
                f"Tu Excel '{filename}' se aplicó correctamente al proyecto {proy.nombre or proy.id}.\nVer: https://herramientas.bigcapital.cl/src/stock-interno/proyecto-vista.html?id={proy.id}",
                _html_ok(proy, filename, res),
            )
            _mark_seen(item.uid)
            procesados.append({"from": item.from_addr, "result": "ok", "proy": proy.id})
    return {"enabled": True, "processed": len(procesados), "items": procesados}
